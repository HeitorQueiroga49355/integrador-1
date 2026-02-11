from django.shortcuts import render, get_object_or_404, redirect
from django.http import HttpResponse
from django.contrib import messages
from django.utils import timezone
from user.models import Profile
from user.utils import get_default_page_alias_by_user
from .forms import ProposalForm
from submission.models import Submission
from django.db.models import Q, Avg
from institution.models import Institution
from .models import Proposal
from datetime import datetime, date, timedelta
from io import BytesIO

# Importações para Excel
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# Importações para PDF
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER


def proposals(request):
    if not request.user.is_authenticated or request.user.profile.role not in [Profile.Role.MANAGER]:
        return redirect(get_default_page_alias_by_user(request.user))
    
    if request.method == 'POST':
        print("=" * 80)
        print("🔵 POST RECEBIDO")
        print("POST DATA:", dict(request.POST))
        print("FILES:", dict(request.FILES))
        print("=" * 80)
        
        form = ProposalForm(request.POST, request.FILES)
        
        print("🔵 VALIDANDO FORMULÁRIO...")
        print("Is Valid:", form.is_valid())
        
        if not form.is_valid():
            print("❌ ERROS DO FORMULÁRIO:")
            for field, errors in form.errors.items():
                print(f"  - Campo '{field}': {errors}")
            print("=" * 80)
            messages.error(request, 'Por favor, corrija os erros do formulário.')
        else:
            print("✅ FORMULÁRIO VÁLIDO")
            try:
                new_proposal = form.save(commit=False)
                institution = Institution.objects.first()
                
                if institution:
                    new_proposal.institution = institution
                    new_proposal.save()
                    print(f"✅ EDITAL SALVO: ID={new_proposal.id}, Título={new_proposal.title}")
                    messages.success(request, f"Edital '{new_proposal.title}' criado com sucesso!")
                    return redirect('proposals:proposals')
                else:
                    print("❌ NENHUMA INSTITUIÇÃO ENCONTRADA")
                    form.add_error(None, "Nenhuma instituição encontrada para vincular o edital.")
            except Exception as e:
                print(f"❌ ERRO AO SALVAR: {e}")
                import traceback
                traceback.print_exc()
                messages.error(request, f'Erro ao salvar: {e}')
    else:
        form = ProposalForm()
    
    today = datetime.now().date()
    editais_abertos = Proposal.objects.filter(closing_date__gte=today).order_by('-opening_date')
    editais_fechados = Proposal.objects.filter(closing_date__lt=today).order_by('-opening_date')
    
    return render(request, 'proposals/proposals.html', {
        'editais_abertos': editais_abertos,
        'editais_fechados': editais_fechados,
        'form': form
    })


def details(request, submission_id):
    if not request.user.is_authenticated or request.user.profile.role not in [Profile.Role.MANAGER]:
        return redirect(get_default_page_alias_by_user(request.user))
    
    submission = get_object_or_404(Submission, id=submission_id)
    context = {
        'submission': submission
    }
    return render(request, 'proposals/details.html', context)


def submissions(request):
    if not request.user.is_authenticated or request.user.profile.role not in [Profile.Role.MANAGER]:
        return redirect(get_default_page_alias_by_user(request.user))
    
    search_query = request.GET.get('q')
    if search_query:
        submissions_list = Submission.objects.filter(
            Q(title__icontains=search_query) |
            Q(researcher__user__first_name__icontains=search_query) |
            Q(researcher__user__username__icontains=search_query)
        )
    else:
        submissions_list = Submission.objects.all().order_by('-created_at')
    
    context = {
        'submissions': submissions_list
    }
    return render(request, 'proposals/submissions.html', context)


def reviewers(request):
    if not request.user.is_authenticated or request.user.profile.role not in [Profile.Role.MANAGER]:
        return redirect(get_default_page_alias_by_user(request.user))
    
    from evaluations.models import Reviewer
    reviewers = Reviewer.objects.all().order_by('-created_at')
    from evaluations.forms import ReviewerForm
    
    context = {
        'reviewers': reviewers,
        'form': ReviewerForm()
    }
    return render(request, 'proposals/reviewers.html', context)


def proposal_edit(request, proposal_id):
    if not request.user.is_authenticated or request.user.profile.role not in [Profile.Role.MANAGER]:
        return redirect(get_default_page_alias_by_user(request.user))
    
    proposal = get_object_or_404(Proposal, id=proposal_id)
    
    if request.method == 'POST':
        form = ProposalForm(request.POST, request.FILES, instance=proposal)
        if form.is_valid():
            form.save()
            messages.success(request, f"Edital '{proposal.title}' atualizado com sucesso!")
            return redirect('proposals:proposals')
    else:
        return redirect('proposals:proposals')


def proposal_delete(request, proposal_id):
    if not request.user.is_authenticated or request.user.profile.role not in [Profile.Role.MANAGER]:
        return redirect(get_default_page_alias_by_user(request.user))
    
    proposal = get_object_or_404(Proposal, id=proposal_id)
    
    if request.method == 'POST':
        title = proposal.title
        proposal.delete()
        messages.success(request, f"Edital '{title}' excluído com sucesso!")
        return redirect('proposals:proposals')


# ==================== NOVAS FUNCIONALIDADES ====================

def close_proposal_manually(request, proposal_id):
    """
    Fecha um edital manualmente e dispara a distribuição automática
    """
    if not request.user.is_authenticated or request.user.profile.role not in [Profile.Role.MANAGER]:
        return redirect(get_default_page_alias_by_user(request.user))
    
    proposal = get_object_or_404(Proposal, id=proposal_id)
    
    if request.method == 'POST':
        # Altera a data de fechamento para ontem (fecha o edital)
        proposal.closing_date = date.today() - timedelta(days=1)
        proposal.save()
        
        messages.success(
            request, 
            f"Edital '{proposal.title}' foi fechado. A distribuição para avaliadores será iniciada automaticamente."
        )
        
        # Redireciona para o status de distribuição
        return redirect('evaluations:distribution_status', proposal_id=proposal_id)
    
    return redirect('proposals:proposals')


def export_proposal_results_excel(request, proposal_id):
    """
    Exporta os resultados de um edital em formato Excel
    """
    if not request.user.is_authenticated or request.user.profile.role not in [Profile.Role.MANAGER]:
        return redirect(get_default_page_alias_by_user(request.user))
    
    from evaluations.models import Evaluation
    
    proposal = get_object_or_404(Proposal, id=proposal_id)
    
    # Cria workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resultados"
    
    # Estilos
    header_fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Cabeçalho do edital
    ws['A1'] = "RELATÓRIO DE AVALIAÇÕES"
    ws['A1'].font = Font(bold=True, size=16)
    ws.merge_cells('A1:I1')
    
    ws['A2'] = f"Edital: {proposal.title}"
    ws['A2'].font = Font(bold=True, size=12)
    ws.merge_cells('A2:I2')
    
    ws['A3'] = f"Período: {proposal.opening_date.strftime('%d/%m/%Y')} a {proposal.closing_date.strftime('%d/%m/%Y')}"
    ws.merge_cells('A3:I3')
    
    ws['A4'] = f"Data do Relatório: {timezone.now().strftime('%d/%m/%Y %H:%M')}"
    ws.merge_cells('A4:I4')
    
    # Cabeçalhos da tabela
    headers = [
        'Posição', 'Título da Submissão', 'Pesquisador', 
        'Nº Avaliações', 'Média Final', 'Relevância Científica', 
        'Viabilidade', 'Resultados Esperados', 'Status'
    ]
    
    row = 6
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Dados das submissões
    submissions = Submission.objects.filter(proposal=proposal)
    report_data = []
    
    for submission in submissions:
        evaluations = Evaluation.objects.filter(
            submission=submission,
            status='completed'
        )
        
        if evaluations.exists():
            avg_score = evaluations.aggregate(Avg('score'))['score__avg']
            avg_relevance = evaluations.aggregate(Avg('note_scientific_relevance'))['note_scientific_relevance__avg']
            avg_viability = evaluations.aggregate(Avg('note_feasibility_methodological'))['note_feasibility_methodological__avg']
            avg_results = evaluations.aggregate(Avg('note_expected_results'))['note_expected_results__avg']
            
            report_data.append({
                'submission': submission,
                'evaluations_count': evaluations.count(),
                'avg_score': avg_score or 0,
                'avg_relevance': avg_relevance or 0,
                'avg_viability': avg_viability or 0,
                'avg_results': avg_results or 0,
                'status': 'Avaliado'
            })
        else:
            report_data.append({
                'submission': submission,
                'evaluations_count': 0,
                'avg_score': 0,
                'avg_relevance': 0,
                'avg_viability': 0,
                'avg_results': 0,
                'status': 'Não Avaliado'
            })
    
    # Ordena por média de pontuação
    report_data.sort(key=lambda x: x['avg_score'], reverse=True)
    
    # Preenche dados
    row = 7
    for position, data in enumerate(report_data, 1):
        ws.cell(row=row, column=1, value=position).border = border
        ws.cell(row=row, column=2, value=data['submission'].title).border = border
        ws.cell(row=row, column=3, value=data['submission'].researcher.user.get_full_name()).border = border
        ws.cell(row=row, column=4, value=data['evaluations_count']).border = border
        
        # Média com formatação
        avg_cell = ws.cell(row=row, column=5, value=round(data['avg_score'], 2))
        avg_cell.border = border
        avg_cell.font = Font(bold=True)
        
        ws.cell(row=row, column=6, value=round(data['avg_relevance'], 2)).border = border
        ws.cell(row=row, column=7, value=round(data['avg_viability'], 2)).border = border
        ws.cell(row=row, column=8, value=round(data['avg_results'], 2)).border = border
        ws.cell(row=row, column=9, value=data['status']).border = border
        
        # Destaca top 3
        if position <= 3 and data['evaluations_count'] > 0:
            for col in range(1, 10):
                ws.cell(row=row, column=col).fill = PatternFill(
                    start_color="D4EDDA", end_color="D4EDDA", fill_type="solid"
                )
        
        row += 1
    
    # Ajusta largura das colunas
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 20
    ws.column_dimensions['I'].width = 15
    
    # Salva em memória
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    # Retorna o arquivo
    response = HttpResponse(
        buffer.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"resultados_{proposal.title.replace(' ', '_')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response


def export_proposal_results_pdf(request, proposal_id):
    """
    Exporta os resultados de um edital em formato PDF
    """
    if not request.user.is_authenticated or request.user.profile.role not in [Profile.Role.MANAGER]:
        return redirect(get_default_page_alias_by_user(request.user))
    
    from evaluations.models import Evaluation
    
    proposal = get_object_or_404(Proposal, id=proposal_id)
    
    # Cria buffer
    buffer = BytesIO()
    
    # Cria documento
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=1*cm,
        leftMargin=1*cm,
        topMargin=2*cm,
        bottomMargin=2*cm
    )
    
    # Estilos
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#1E40AF'),
        spaceAfter=12,
        alignment=TA_CENTER
    )
    
    subtitle_style = ParagraphStyle(
        'CustomSubtitle',
        parent=styles['Normal'],
        fontSize=12,
        spaceAfter=20,
        alignment=TA_CENTER
    )
    
    # Elementos do PDF
    elements = []
    
    # Título
    elements.append(Paragraph("RELATÓRIO DE AVALIAÇÕES", title_style))
    elements.append(Paragraph(f"<b>Edital:</b> {proposal.title}", subtitle_style))
    elements.append(Paragraph(
        f"<b>Período:</b> {proposal.opening_date.strftime('%d/%m/%Y')} a {proposal.closing_date.strftime('%d/%m/%Y')}",
        subtitle_style
    ))
    elements.append(Spacer(1, 0.5*cm))
    
    # Dados
    submissions = Submission.objects.filter(proposal=proposal)
    report_data = []
    
    for submission in submissions:
        evaluations = Evaluation.objects.filter(
            submission=submission,
            status='completed'
        )
        
        if evaluations.exists():
            avg_score = evaluations.aggregate(Avg('score'))['score__avg']
            report_data.append({
                'submission': submission,
                'avg_score': avg_score or 0,
                'count': evaluations.count()
            })
    
    # Ordena por média
    report_data.sort(key=lambda x: x['avg_score'], reverse=True)
    
    # Tabela
    table_data = [[
        'Pos.', 'Título da Submissão', 'Pesquisador', 
        'Nº Aval.', 'Média Final'
    ]]
    
    for position, data in enumerate(report_data, 1):
        title_text = data['submission'].title[:50] + '...' if len(data['submission'].title) > 50 else data['submission'].title
        table_data.append([
            str(position),
            title_text,
            data['submission'].researcher.user.get_full_name(),
            str(data['count']),
            f"{data['avg_score']:.2f}"
        ])
    
    # Cria tabela
    table = Table(table_data, colWidths=[2*cm, 12*cm, 7*cm, 3*cm, 3*cm])
    
    # Estilo da tabela
    table.setStyle(TableStyle([
        # Cabeçalho
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E40AF')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        
        # Corpo
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('ALIGN', (1, 1), (2, -1), 'LEFT'),
        
        # Grid
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        
        # Top 3
        ('BACKGROUND', (0, 1), (-1, min(3, len(table_data)-1)), colors.HexColor('#D4EDDA')),
    ]))
    
    elements.append(table)
    
    # Gera PDF
    doc.build(elements)
    buffer.seek(0)
    
    # Retorna
    response = HttpResponse(buffer.read(), content_type='application/pdf')
    filename = f"resultados_{proposal.title.replace(' ', '_')}.pdf"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response